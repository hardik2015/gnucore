"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020,2019 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@gmail.com>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Bhavesh Bawadhane" <bbhavesh07@gmail.com>
"Prajkta Patkar" <prajkta.patkar007@gmail.com>
"""


from pyramid.view import view_defaults,  view_config
from gkcore.views.api_login import authCheck
from gkcore.views.api_tax import calTax
from gkcore import eng, enumdict
from pyramid.request import Request
from pyramid.response import Response
from gkcore.models import gkdb
from sqlalchemy.sql import select, distinct
from sqlalchemy import func, desc
import json
from sqlalchemy.engine.base import Connection
from sqlalchemy import and_, exc, func
import jwt
import gkcore
from gkcore.models.meta import dbconnect
from gkcore.models.gkdb import goprod, product,accounts, organisation
from gkcore.views.api_user import getUserRole
from gkcore.views.api_godown import getusergodowns
from datetime import datetime,date
from time import strftime, strptime
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import base64
from io import BytesIO

@view_defaults(route_name='products')
class api_product(object):
    def __init__(self,request):
        self.request = Request
        self.request = request
        self.con = Connection



    @view_config(request_method='GET', renderer ='json')
    def getAllProducts(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con=eng.connect()
                userrole = getUserRole(authDetails["userid"])
                gorole = userrole["gkresult"]
                if (gorole["userrole"]==3):
                    uId = getusergodowns(authDetails["userid"])
                    gid=[]
                    for record1 in uId["gkresult"]:
                        gid.append(record1["goid"])
                    productCodes=[]
                    for record2 in gid:
                        proCode = self.con.execute(select([gkdb.goprod.c.productcode]).where(gkdb.goprod.c.goid==record2))
                        proCodes = proCode.fetchall()
                        for record3 in proCodes:
                            if record3["productcode"] not in productCodes:
                                productCodes.append(record3["productcode"])
                    results = []
                    for record4 in productCodes:
                        result = self.con.execute(select([gkdb.product.c.productcode, gkdb.product.c.productdesc, gkdb.product.c.categorycode, gkdb.product.c.uomid,gkdb.product.c.gsflag, gkdb.product.c.prodsp,gkdb.product.c.prodmrp]).where(and_(gkdb.product.c.orgcode==authDetails["orgcode"], gkdb.product.c.productcode==record4)).order_by(gkdb.product.c.productdesc))
                        products = result.fetchone()
                        results.append(products)
                else:
                    invdc = 9
                    try:
                        invdc = int(self.request.params["invdc"])
                    except:
                        invdc = 9
                    if invdc == 4:
                        results = self.con.execute(select([gkdb.product.c.productcode,gkdb.product.c.gsflag ,gkdb.product.c.productdesc, gkdb.product.c.categorycode, gkdb.product.c.uomid,gkdb.product.c.prodsp,gkdb.product.c.prodmrp]).where(and_(gkdb.product.c.orgcode==authDetails["orgcode"],gkdb.product.c.gsflag==7)).order_by(gkdb.product.c.productdesc))
                    if invdc == 9:
                        results = self.con.execute(select([gkdb.product.c.productcode, gkdb.product.c.productdesc,gkdb.product.c.gsflag, gkdb.product.c.categorycode, gkdb.product.c.uomid,gkdb.product.c.prodsp,gkdb.product.c.prodmrp]).where(gkdb.product.c.orgcode==authDetails["orgcode"]).order_by(gkdb.product.c.productdesc))

                products = []
                srno=1
                for row in results:
                    unitsofmeasurement = self.con.execute(select([gkdb.unitofmeasurement.c.unitname]).where(gkdb.unitofmeasurement.c.uomid==row["uomid"]))
                    unitofmeasurement = unitsofmeasurement.fetchone()
                    if unitofmeasurement != None:
                        unitname = unitofmeasurement["unitname"]
                    else:
                        unitname = ""
                    if row["categorycode"]!=None:
                        categories = self.con.execute(select([gkdb.categorysubcategories.c.categoryname]).where(gkdb.categorysubcategories.c.categorycode==row["categorycode"]))
                        category = categories.fetchone()
                        categoryname = category["categoryname"]
                    else:
                        categoryname=""
                    if row["productcode"]!=None:
                        openingStockResult = self.con.execute(select([gkdb.product.c.openingstock]).where(gkdb.product.c.productcode == row["productcode"]))
                        osRow =openingStockResult.fetchone()
                        openingStock = osRow["openingstock"]
                        productstockin = self.con.execute(select([func.sum(gkdb.stock.c.qty).label("sumofins")]).where(and_(gkdb.stock.c.productcode==row["productcode"],gkdb.stock.c.inout==9)))
                        stockinsum = productstockin.fetchone()
                        if stockinsum["sumofins"]!=None:
                            openingStock = openingStock + stockinsum["sumofins"]
                        productstockout = self.con.execute(select([func.sum(gkdb.stock.c.qty).label("sumofouts")]).where(and_(gkdb.stock.c.productcode==row["productcode"],gkdb.stock.c.inout==15)))
                        stockoutsum = productstockout.fetchone()
                        if stockoutsum["sumofouts"]!=None:
                            openingStock = openingStock - stockoutsum["sumofouts"]
                    products.append({"srno":srno, "unitname":unitname, "categoryname":categoryname, "productcode": row["productcode"], "productdesc":row["productdesc"] , "categorycode": row["categorycode"], "productquantity": "%.2f"%float(openingStock),"gsflag":row["gsflag"]})
                    srno = srno+1
                return {"gkstatus":enumdict["Success"], "gkresult":products}
            except:
                self.con.close()
                return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()


    @view_config(request_param='qty=single', request_method='GET',renderer='json')
    def getProduct(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                result = self.con.execute(select([gkdb.product]).where(gkdb.product.c.productcode==self.request.params["productcode"]))
                row = result.fetchone()
                productDetails={ "discountpercent":"%.2f"%float(row["percentdiscount"]),"discountamount":"%.2f"%float(row["amountdiscount"]),"productcode":row["productcode"],"productdesc": row["productdesc"], "gsflag":row["gsflag"],"gscode":row["gscode"]}

                # the field deletable is for check whether product/service are in use or not
                #first it check that product/service are use in stock table and purchaseorder table and then give count of product/service are in use
                #if count is grater than 0 it send 1 else it send 0 as value of deletable key
                if int(row["gsflag"]) == 19:
                    prod_countinv =self.con.execute("SELECT (contents ::json)->'%s' is NULL FROM invoice where orgcode ='%d'"%((str(self.request.params["productcode"])),(int(authDetails["orgcode"])))).fetchall()
                    if (False,) in prod_countinv:
                        productDetails["deletable"] = 1
                    else:
                        prod_purch =self.con.execute("SELECT (schedule ::json)->'%s' is NULL FROM purchaseorder where orgcode ='%d'"%((str(self.request.params["productcode"])),(int(authDetails["orgcode"])))).fetchall()
                        if (False,) in prod_purch:
                            productDetails["deletable"] = 1
                    
                if row["prodsp"]!=None:
                    productDetails["prodsp"] = "%.2f"%float(row["prodsp"])
                else:
                    productDetails["prodsp"] = "%.2f"%0.00
                if row["prodmrp"]!=None:
                    productDetails["prodmrp"] = "%.2f"%float(row["prodmrp"])
                else:
                    productDetails["prodmrp"] = "%.2f"%0.00
                if int(row["gsflag"]) == 7:
                    prod_countinstock = self.con.execute("select count(productcode) as pccount from stock where productcode='%s' and orgcode='%d'"%((str(self.request.params["productcode"])),(int(authDetails["orgcode"]))))
                    pc_countinstock = prod_countinstock.fetchone()

                    if pc_countinstock["pccount"] > 0:
                        productDetails["deletable"] = 1

                    else: 
                        prod_countinpuchaseorder = self.con.execute("select count(purchaseorder.schedule) as pccount from purchaseorder where purchaseorder.schedule?'%s'and orgcode='%d'"%((str(self.request.params["productcode"])),(int(authDetails["orgcode"]))))
                        pc_countinpuchaseorder = prod_countinpuchaseorder.fetchone()
                        if pc_countinpuchaseorder["pccount"] > 0:
                            productDetails["deletable"] = 1  
                        else:
                            productDetails["deletable"] = 0
                    result1 = self.con.execute(select([gkdb.unitofmeasurement.c.unitname]).where(gkdb.unitofmeasurement.c.uomid==row["uomid"]))
                    unitrow= result1.fetchone()
                    productDetails["specs"] = row["specs"]
                    productDetails["categorycode"] = row["categorycode"]
                    productDetails["uomid"]=row["uomid"]
                    productDetails["gsflag"]=row["gsflag"]
                    productDetails["unitname"]=unitrow["unitname"]
                    productDetails["openingstock"]="%.2f"%float(row["openingstock"])
                    userrole = getUserRole(authDetails["userid"])
                    if int(userrole["gkresult"]["userrole"]) != 3:
                        godownswithstock = self.con.execute(select([func.count(gkdb.goprod.c.productcode).label("numberofgodowns")]).where(gkdb.goprod.c.productcode==self.request.params["productcode"]))
                        godowns = godownswithstock.fetchone()
                        numberofgodowns = godowns["numberofgodowns"]
                    else:
                        usergodowmns = getusergodowns(authDetails["userid"])
                        numberofgodowns = 0
                        for usergodown in usergodowmns["gkresult"]:
                            godownswithstock = self.con.execute(select([gkdb.goprod.c.goid]).where(and_(gkdb.goprod.c.productcode==self.request.params["productcode"], gkdb.goprod.c.goid==usergodown["goid"])))
                            usergodownwithstock = godownswithstock.fetchone()
                            try:
                                if usergodownwithstock['goid']:
                                    numberofgodowns = numberofgodowns + 1
                            except:
                                continue
                    return {"gkstatus":enumdict["Success"],"gkresult":productDetails,"numberofgodowns":"%d"%int(numberofgodowns)}
                else:
                    return {"gkstatus":enumdict["Success"],"gkresult":productDetails}

            except:
                self.con.close()
                return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()
    @view_config(request_method='GET',request_param='type=pt',renderer='json')
    def getTaxForProduct(self):
        """
        Purpose: returns either VAT or GST for a selected product based on product code and state.
        description:
        This function takes productcode,source and destination states,
        (called source and destination as params).
        Also takes taxflag.
        The function makes calld to the global function calTax found in api_tax.
        Will return a dictionary containing the tax name and rate.
        Please refer calTax in api_tax for details.
        """
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                return calTax(int(self.request.params["taxflag"]),self.request.params["source"],self.request.params["destination"],int(self.request.params["productcode"]),self.con)

            except:
                self.con.close()
                return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()





    @view_config(request_method='GET', request_param='by=category',renderer='json')
    def getProductbyCategory(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                if self.request.params["categorycode"] =="":
                    result = self.con.execute(select([gkdb.product.c.productcode,gkdb.product.c.productdesc]).where(gkdb.product.c.categorycode==None))
                else:
                    result = self.con.execute(select([gkdb.product.c.productcode,gkdb.product.c.productdesc]).where(gkdb.product.c.categorycode==self.request.params["categorycode"]))
                prodlist = []
                for row in result:
                    productDetails={ "productcode":row["productcode"],"productdesc": row["productdesc"]}
                    prodlist.append(productDetails);
                return {"gkstatus":enumdict["Success"],"gkresult":prodlist}
            except:
                self.con.close()
                return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()


    @view_config(request_method='GET', request_param='by=godown',renderer='json')
    def getProductbyGodown(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                productcode = self.request.params["productcode"]
                userrole = getUserRole(authDetails["userid"])
                if int(userrole["gkresult"]["userrole"]) != 3:
                    result = self.con.execute(select([goprod]).where(goprod.c.productcode == productcode))
                    godowns = []
                    for row in result:
                        goDownDetails = {"goid":row["goid"], "goopeningstock":"%.2f"%float(row["goopeningstock"]), "productcode":row["productcode"]}
                        godowns.append(goDownDetails)
                else:
                    usergodowns = getusergodowns(authDetails["userid"])
                    godowns = []
                    for usergodown in usergodowns["gkresult"]:
                        thisgodown = self.con.execute(select([goprod]).where(and_(goprod.c.productcode == productcode, goprod.c.goid == usergodown["goid"])))
                        thisgodown = thisgodown.fetchone()
                        try:
                            goDownDetails = {"goid":thisgodown["goid"], "goopeningstock":"%.2f"%float(thisgodown["goopeningstock"]), "productcode":thisgodown["productcode"]}
                            godowns.append(goDownDetails)
                        except:
                            continue
                return {"gkstatus":enumdict["Success"],"gkresult":godowns}
            except:
                self.con.close()
                return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()

    @view_config(request_method='GET', request_param='from=godown',renderer='json')
    def getProductfromGodown(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                goid = self.request.params["godownid"]
                result = self.con.execute(select([gkdb.goprod.c.goprodid, gkdb.goprod.c.goopeningstock, gkdb.goprod.c.productcode]).where(and_(gkdb.goprod.c.goid== goid, gkdb.goprod.c.orgcode==authDetails["orgcode"])))
                products = []
                for row in result:
                    productDetails = {"goprodid":row["goprodid"], "goopeningstock":"%.2f"%float(row["goopeningstock"]), "productcode":row["productcode"]}
                    products.append(productDetails)
                return {"gkstatus":enumdict["Success"],"gkresult":products}
            except:
                self.con.close()
                return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()


    @view_config(request_method='POST',renderer='json')
    def addProduct(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                dataset = self.request.json_body
                productDetails = dataset["productdetails"]
                godownFlag = dataset["godownflag"]
                productDetails["orgcode"] = authDetails["orgcode"]
                if ("categorycode" in productDetails)==False:
                    duplicateproduct = self.con.execute(select([func.count(gkdb.product.c.productcode).label("productcount")]).where(and_(gkdb.product.c.productdesc== productDetails["productdesc"],gkdb.product.c.categorycode==None,gkdb.product.c.orgcode==productDetails["orgcode"])))
                    duplicateproductrow = duplicateproduct.fetchone()
                    if duplicateproductrow["productcount"]>0:
                        return {"gkstatus":enumdict["DuplicateEntry"]}
                result = self.con.execute(gkdb.product.insert(),[productDetails])
                spec = productDetails["specs"]
                for sp in list(spec.keys()):
                    self.con.execute("update categoryspecs set productcount = productcount +1 where spcode = %d"%(int(sp)))
                if ("categorycode" in productDetails)==False:
                    productDetails["categorycode"]=None
                result = self.con.execute(select([gkdb.product.c.productcode]).where(and_(gkdb.product.c.productdesc==productDetails["productdesc"], gkdb.product.c.categorycode==productDetails["categorycode"],gkdb.product.c.orgcode==productDetails["orgcode"])))
                row = result.fetchone()
                productCode = row["productcode"]
                if godownFlag:
                    goDetails = dataset["godetails"]
                    ttlOpening = 0.00
                    for g in list(goDetails.keys()):
                        ttlOpening = ttlOpening + float(goDetails[g])
                        goro = {"productcode":productCode,"goid":g,"goopeningstock":goDetails[g],"orgcode":authDetails["orgcode"]}
                        self.con.execute(goprod.insert(),[goro])
                    self.con.execute(product.update().where(and_(product.c.productcode == productCode,product.c.orgcode==authDetails["orgcode"])).values(openingstock = ttlOpening))

                # We need to create sale and purchase accounts for product under sales and purchase groups respectively.
                sp = self.con.execute("select groupcode from groupsubgroups where groupname in ('%s','%s') and orgcode = %d"%('Sales','Purchase',productDetails["orgcode"]))
                s = sp.fetchall()
                prodName = productDetails["productdesc"]
                proSale = prodName + " Sale"
                proPurch = prodName + " Purchase"
                resultb = self.con.execute(gkdb.accounts.insert(),[{"accountname":proPurch,"groupcode":s[0][0],"orgcode":authDetails["orgcode"],"sysaccount":1},{"accountname":proSale,"groupcode":s[1][0],"orgcode":authDetails["orgcode"],"sysaccount":1}])
                
                return {"gkstatus":enumdict["Success"],"gkresult":row["productcode"]}

            except exc.IntegrityError:
                return {"gkstatus":enumdict["DuplicateEntry"]}
            except:
                return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()

    '''
    Here product data is updated with new data input by the user while editing product.
    If godowns have been created and godownwise opening stock has been entered for a product the record in "goprod" table is first deleted and fresh record is created.
    If godownwise opening stock was not recorded while creating product it can be created here.
    In this case a new record is created in "goprod" table.
    '''
    @view_config(request_method='PUT', renderer='json')
    def editProduct(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                dataset = self.request.json_body
                productDetails = dataset["productdetails"]
                
                godownFlag = dataset["godownflag"]
                productCode = productDetails["productcode"]
                pn = self.con.execute(select([gkdb.product.c.productdesc]).where(gkdb.product.c.productcode==productCode))
                prodName = pn.fetchone()
                result = self.con.execute(gkdb.product.update().where(gkdb.product.c.productcode==productDetails["productcode"]).values(productDetails))
                if godownFlag:
                    goDetails = dataset["godetails"]
                    result = self.con.execute(gkdb.goprod.delete().where(and_(gkdb.goprod.c.productcode==productCode,gkdb.goprod.c.orgcode==authDetails["orgcode"])))
                    ttlOpening = 0.0
                    for g in list(goDetails.keys()):
                        ttlOpening = ttlOpening + float(goDetails[g])
                        goro = {"productcode":productCode,"goid":g,"goopeningstock":goDetails[g],"orgcode":authDetails["orgcode"]}
                        self.con.execute(gkdb.goprod.insert(),[goro])
                    self.con.execute(product.update().where(and_(product.c.productcode == productCode,product.c.orgcode==authDetails["orgcode"])).values(openingstock = ttlOpening))
                # We need to update accountname also.
                pnSL = str(prodName["productdesc"]) +" Sale"
                newpnSL = str(productDetails["productdesc"])+" Sale"
                pnPurch = str(prodName["productdesc"]) +" Purchase"
                newpnPH = str(productDetails["productdesc"])+" Purchase"
                self.con.execute(accounts.update().where(and_(accounts.c.accountname == pnSL,accounts.c.orgcode == authDetails["orgcode"])).values(accountname = newpnSL))
                self.con.execute(accounts.update().where(and_(accounts.c.accountname == pnPurch,accounts.c.orgcode == authDetails["orgcode"])).values(accountname = newpnPH))
                return {"gkstatus":enumdict["Success"]}
            except exc.IntegrityError:
                return {"gkstatus":enumdict["DuplicateEntry"]}
            except:
                return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()

    @view_config(request_method='DELETE', renderer ='json')
    def deleteProduct(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                dataset = self.request.json_body
                result = self.con.execute(select([gkdb.product.c.specs,gkdb.product.c.productdesc]).where(gkdb.product.c.productcode==dataset["productcode"]))
                row = result.fetchone()
                spec = row["specs"]
                pn = row["productdesc"]
                for sp in list(spec.keys()):
                    self.con.execute("update categoryspecs set productcount = productcount -1 where spcode = %d"%(int(sp)))
                
                result = self.con.execute(gkdb.product.delete().where(gkdb.product.c.productcode==dataset["productcode"]))
                try:
                    resultA = self.con.execute(accounts.delete().where(and_(accounts.c.accountname.like(pn+'%'),accounts.c.orgcode == authDetails["orgcode"])))
                except:
                    pass
                return {"gkstatus":enumdict["Success"]}
            except exc.IntegrityError:
                return {"gkstatus":enumdict["ActionDisallowed"]}
            except:
                return {"gkstatus":enumdict["ConnectionFailed"] }
            finally:
                self.con.close()

    @view_config(request_method='GET',request_param='tax=vatorgst',renderer='json')
    def getvatorgst(self):
        """
        Purpose:
        To determine what kind of tax will be applible to the goods of corresponding organisation.
        Description:
        This function uses orgcode of organisation and will fetch it's financialStart and financialEnddate.
        gstdate as "01/07/2017" is used which is compared with financialStart and financialEnd based on this gstorvatflag changes as following:
        gstorvatflag=7 (means GST is applible)
        gstorvatflag=22 (means VAT is applible)
        gstorvatflag=29 (GST and VAT are applible)
        """
        try:
            token=self.request.headers["gktoken"]
        except :
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con=eng.connect()
                result = self.con.execute(select([gkdb.organisation.c.yearstart,gkdb.organisation.c.yearend]).where(gkdb.organisation.c.orgcode==authDetails["orgcode"]))
                yearstartandend=result.fetchone()
                gstorvatflag=29
                date1 = "2017-07-01"
                gstdate = datetime.strptime(date1, "%Y-%m-%d")
                financialStart = datetime.strptime(str(yearstartandend["yearstart"]), "%Y-%m-%d")
                financialEnd = datetime.strptime(str(yearstartandend["yearend"]),"%Y-%m-%d")

                if (gstdate>financialStart and gstdate>financialEnd):
                    gstorvatflag=22
                elif(gstdate>financialStart and gstdate<=financialEnd) :
                    gstorvatflag=29
                elif(gstdate<=financialStart and gstdate<=financialEnd):
                    gstorvatflag=7
                return {"gkstatus":enumdict["Success"],"gkresult":str(gstorvatflag)}
            except:
                return {"gkstatus":enumdict["ConnectionFailed"] }
            finally:
                self.con.close()

    '''
    A godown keeper can only access the list of products that are present in the godowns assigned to him.
    This function lets a godown keeper access the list of all products in an organisation.
    Also, godown incharge cannot access products which are already having openingStock for particular godown which is selected.
    '''

    @view_config(request_method='GET', request_param='list=allprod', renderer ='json')
    def getAllProdList(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con=eng.connect()
                currentgoid=int(self.request.params["goid"])
                userrole = getUserRole(authDetails["userid"])
                gorole = userrole["gkresult"]
                if gorole["userrole"]==3:
                    uId = getusergodowns(authDetails["userid"])
                    gid=[]
                    for record1 in uId["gkresult"]:
                        gid.append(record1["goid"])
                    if currentgoid in gid:
                        proCode = self.con.execute(select([gkdb.goprod.c.productcode]).where(gkdb.goprod.c.goid==currentgoid))
                        proCodes = proCode.fetchall()
                    productCodes=[]                     
                    for record3 in proCodes:
                        if record3["productcode"] not in productCodes:
                            productCodes.append(record3["productcode"])
                    results = self.con.execute(select([gkdb.product.c.productcode,gkdb.product.c.productdesc]).where(and_(gkdb.product.c.orgcode==authDetails["orgcode"],gkdb.product.c.gsflag==7)).order_by(gkdb.product.c.productdesc))
                    products = []
                    for row in results:
                        if row["productcode"] not in productCodes:
                            products.append({"productcode": row["productcode"], "productdesc":row["productdesc"]})
                    
                    return {"gkstatus":enumdict["Success"], "gkresult":products}
            except:
               self.con.close()
               return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()

    '''
    This function is written for fetching the HSN code, UOM automatically when product is selected.
    '''
    @view_config(request_method='GET', request_param='type=hsnuom', renderer ='json')
    def gethsnuom(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con=eng.connect()
                product = self.con.execute(select([gkdb.product.c.uomid,gkdb.product.c.gscode]).where(and_(gkdb.product.c.productcode==self.request.params["productcode"],gkdb.product.c.orgcode==authDetails["orgcode"])))
                productdetails = product.fetchone()
                uom = self.con.execute(select([gkdb.unitofmeasurement.c.unitname]).where(gkdb.unitofmeasurement.c.uomid==productdetails["uomid"]))
                unitname = uom.fetchone()
                productDetails={"unitname": unitname["unitname"],"gscode":productdetails["gscode"]}
                return {"gkstatus":enumdict["Success"], "gkresult":productDetails}
            except:
                self.con.close()
                return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()

    '''
    This is a function for saving opening stock for the selected product
    ''' 
    @view_config(request_method='POST', request_param='type=addstock', renderer='json')
    def addstock(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
             try:
                self.con = eng.connect()
                dataset = self.request.json_body
                orgcode=authDetails["orgcode"]
                goid=dataset["goid"]
                productDetails=dataset["productdetails"]
                for product in productDetails:
                    details={"goid":goid,"goopeningstock":productDetails[product],"productcode":product,"orgcode":orgcode}
                    result=self.con.execute(gkdb.goprod.insert(),[details])
                return {"gkstatus":enumdict["Success"]}
             except exc.IntegrityError:
                return {"gkstatus":enumdict["DuplicateEntry"]}
             except:
                return {"gkstatus":enumdict["ConnectionFailed"]}
             finally:
                self.con.close()


    '''
    This funtion returns the last price for which a product was sold/purchased to/from a party.
    To find out the price the function needs productcode of the product, custid of the party and inoutflag(to determine whether selling/purchase price is needed)
    A select query first fetches the invid of the last sale/purchase invoice created for the party involving the product.
    Another query retrives the contents of the invoice whose invid is same as the result of the above query.
    Price is found out from the contents using productcode as key and sent as response.
    '''
    @view_config(request_param='type=lastprice', request_method='GET',renderer='json')
    def lastPrice(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                lastPriceData = self.con.execute(select([gkdb.cslastprice.c.lastprice]).where(and_(gkdb.cslastprice.c.custid==int(self.request.params["custid"]), gkdb.cslastprice.c.productcode==int(self.request.params["productcode"]), gkdb.cslastprice.c.inoutflag==int(self.request.params["inoutflag"]), gkdb.cslastprice.c.orgcode==int(authDetails["orgcode"]))))
                lastPriceValue = lastPriceData.fetchone()["lastprice"]
                return {"gkstatus":enumdict["Success"], "gkresult":"%.2f"%float(lastPriceValue)}
            except:
                    return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()

    '''
    This function returns a list of products and services in the form of a spreadsheet.
    '''
    @view_config(request_method='GET',request_param="type=spreadsheet", renderer ='json')
    def getListofProductsSpreadsheet(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  gkcore.enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"] == False:
            return  {"gkstatus":  gkcore.enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con=eng.connect()
                orgdetails = self.con.execute(select([gkdb.organisation.c.orgname,gkdb.organisation.c.yearstart,gkdb.organisation.c.yearend]).where(gkdb.organisation.c.orgcode==authDetails["orgcode"]))
                orgData=orgdetails.fetchone()
                gstorvatflag=29
                date1 = "2017-07-01"
                gstdate = datetime.strptime(date1, "%Y-%m-%d")
                financialStart = datetime.strptime(str(orgData["yearstart"]), "%Y-%m-%d")
                financialEnd = datetime.strptime(str(orgData["yearend"]),"%Y-%m-%d")
                fystart = financialStart.strftime("%d-%m-%Y")
                fyend = financialEnd.strftime("%d-%m-%Y")
                orgname = orgData["orgname"]

                if (gstdate>financialStart and gstdate>financialEnd):
                    gstorvatflag=22
                elif(gstdate>financialStart and gstdate<=financialEnd) :
                    gstorvatflag=29
                elif(gstdate<=financialStart and gstdate<=financialEnd):
                    gstorvatflag=7
                userrole = getUserRole(authDetails["userid"])
                gorole = userrole["gkresult"]
                if (gorole["userrole"]==3):
                    uId = getusergodowns(authDetails["userid"])
                    gid=[]
                    for record1 in uId["gkresult"]:
                        gid.append(record1["goid"])
                    productCodes=[]
                    for record2 in gid:
                        proCode = self.con.execute(select([gkdb.goprod.c.productcode]).where(gkdb.goprod.c.goid==record2))
                        proCodes = proCode.fetchall()
                        for record3 in proCodes:
                            if record3["productcode"] not in productCodes:
                                productCodes.append(record3["productcode"])
                    results = []
                    for record4 in productCodes:
                        result = self.con.execute(select([gkdb.product.c.productcode, gkdb.product.c.productdesc, gkdb.product.c.categorycode, gkdb.product.c.uomid,gkdb.product.c.gsflag, gkdb.product.c.prodsp,gkdb.product.c.prodmrp]).where(and_(gkdb.product.c.orgcode==authDetails["orgcode"], gkdb.product.c.productcode==record4)).order_by(gkdb.product.c.productdesc))
                        products = result.fetchone()
                        results.append(products)
                else:
                    invdc = 9
                    try:
                        invdc = int(self.request.params["invdc"])
                    except:
                        invdc = 9
                    if invdc == 4:
                        results = self.con.execute(select([gkdb.product.c.productcode,gkdb.product.c.gsflag ,gkdb.product.c.productdesc, gkdb.product.c.categorycode, gkdb.product.c.uomid,gkdb.product.c.prodsp,gkdb.product.c.prodmrp]).where(and_(gkdb.product.c.orgcode==authDetails["orgcode"],gkdb.product.c.gsflag==7)).order_by(gkdb.product.c.productdesc))
                    if invdc == 9:
                        results = self.con.execute(select([gkdb.product.c.productcode, gkdb.product.c.productdesc,gkdb.product.c.gsflag, gkdb.product.c.categorycode, gkdb.product.c.uomid,gkdb.product.c.prodsp,gkdb.product.c.prodmrp]).where(gkdb.product.c.orgcode==authDetails["orgcode"]).order_by(gkdb.product.c.productdesc))

                products = []
                srno=1
                for productrow in results:
                    unitsofmeasurement = self.con.execute(select([gkdb.unitofmeasurement.c.unitname]).where(gkdb.unitofmeasurement.c.uomid==productrow["uomid"]))
                    unitofmeasurement = unitsofmeasurement.fetchone()
                    if unitofmeasurement != None:
                        unitname = unitofmeasurement["unitname"]
                    else:
                        unitname = ""
                    if productrow["categorycode"]!=None:
                        categories = self.con.execute(select([gkdb.categorysubcategories.c.categoryname]).where(gkdb.categorysubcategories.c.categorycode==productrow["categorycode"]))
                        category = categories.fetchone()
                        categoryname = category["categoryname"]
                    else:
                        categoryname=""
                    if productrow["productcode"]!=None:
                        openingStockResult = self.con.execute(select([gkdb.product.c.openingstock]).where(gkdb.product.c.productcode == productrow["productcode"]))
                        osProductrow =openingStockResult.fetchone()
                        openingStock = osProductrow["openingstock"]
                        productstockin = self.con.execute(select([func.sum(gkdb.stock.c.qty).label("sumofins")]).where(and_(gkdb.stock.c.productcode==productrow["productcode"],gkdb.stock.c.inout==9)))
                        stockinsum = productstockin.fetchone()
                        if stockinsum["sumofins"]!=None:
                            openingStock = openingStock + stockinsum["sumofins"]
                        productstockout = self.con.execute(select([func.sum(gkdb.stock.c.qty).label("sumofouts")]).where(and_(gkdb.stock.c.productcode==productrow["productcode"],gkdb.stock.c.inout==15)))
                        stockoutsum = productstockout.fetchone()
                        if stockoutsum["sumofouts"]!=None:
                            openingStock = openingStock - stockoutsum["sumofouts"]
                    products.append({"srno":srno, "unitname":unitname, "categoryname":categoryname, "productcode": productrow["productcode"], "productdesc":productrow["productdesc"] , "categorycode": productrow["categorycode"], "productquantity": "%.2f"%float(openingStock),"gsflag":productrow["gsflag"]})
                    srno = srno+1
                # A workbook is opened.
                productwb = openpyxl.Workbook()
                # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
                sheet = productwb.active
                # Title of the sheet and width of columns are set.
                sheet.title = "List of Products"
                sheet.column_dimensions['A'].width = 8
                sheet.column_dimensions['B'].width = 24
                sheet.column_dimensions['C'].width = 18
                sheet.column_dimensions['D'].width = 24
                sheet.column_dimensions['E'].width = 16
                # Cells of first two rows are merged to display organisation details properly.
                sheet.merge_cells('A1:E2')
                # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
                sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
                sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
                # Organisation name and financial year are displayed.
                sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
                sheet.merge_cells('A3:E3')
                sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
                sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
                sheet['A3'] = 'List of Products'
                sheet.merge_cells('A3:E3')
                sheet['A4'] = 'Sr.No.'
                if gstorvatflag == "22":
                    sheet['B4'] = 'Product'
                    sheet['C4'] = 'Category'
                    sheet['D4'] = 'UOM'
                else:
                    sheet['B4'] = 'Product/service'
                    sheet['C4'] = 'Type'
                    sheet['D4'] = 'Category'
                    sheet['E4'] = 'Uom'
                titlerow = sheet.row_dimensions[4]
                titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
                srno=1
                if gstorvatflag == "22":
                    row = 5
                    for stock in products:
                        sheet['A'+str(row)] = srno
                        sheet['A'+str(row)].alignment = Alignment(horizontal='left')
                        sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                        sheet['B'+str(row)] = stock['productdesc']
                        sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                        sheet['C'+str(row)] = stock["categoryname"]
                        sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                        sheet['D'+str(row)] = stock["unitname"]
                        sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                        row +=1
                        srno +=1
                else:
                    row = 5
                    for stock in products:
                        sheet['A'+str(row)] = srno
                        sheet['A'+str(row)].alignment = Alignment(horizontal='left')
                        sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                        sheet['B'+str(row)] = stock['productdesc']
                        sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                        if stock["gsflag"] == 7:
                                  sheet['C'+str(row)] = 'Product'
                                  sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                        else:    
                                 sheet['C'+str(row)] = 'Service'
                                 sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                        sheet['D'+str(row)] = stock["categoryname"]
                        sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                        sheet['E'+str(row)] = stock["unitname"]
                        sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                        row +=1
                        srno +=1
                output = io.BytesIO()
                productwb.save(output)
                contents = output.getvalue()
                output.close()
                headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','X-Content-Type-Options':'nosniff', 'Set-Cookie':'fileDownload=true; path=/ [;HttpOnly]'}
                # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/'}
                return Response(contents, headerlist=list(headerList.items()))
            except:
               self.con.close()
               return {"gkstatus":enumdict["ConnectionFailed"]}
            finally:
                self.con.close()
